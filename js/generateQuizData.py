#!/usr/bin/env python3
"""
Generador de preguntas de quiz para Career Quest
Crea preguntas basadas en funciones de puestos en Excel
"""

import openpyxl
import json
import random
import re
import unicodedata
from collections import defaultdict

def slugify(text):
    """Convierte texto a slug (lowercase con hyphens)"""
    text = unicodedata.normalize('NFKD', str(text))
    text = text.encode('ASCII', 'ignore').decode('ASCII')
    text = re.sub(r'[^\w\s-]', '', text).strip()
    text = re.sub(r'[-\s]+', '-', text)
    return text.lower()

# Cargar Excel
wb = openpyxl.load_workbook('js/EXCEL/Informacion - Puestos LyC VF.xlsx')
ws_puestos = wb['Lista de Puestos']
ws_funciones = wb['Funciones']

# Extraer funciones del Excel
funciones_dict = defaultdict(list)
funciones_por_area = defaultdict(list)

for row in ws_funciones.iter_rows(min_row=2, max_row=ws_funciones.max_row, values_only=True):
    if row[0] and row[1]:
        puesto_name = row[0].strip()
        funcion = row[1].strip()
        if funcion and len(funcion) > 10:  # Filtrar funciones válidas
            funciones_dict[puesto_name].append(funcion)

# Extraer info de puestos
areas_data = defaultdict(lambda: {
    'positions': [],
    'all_functions': [],
    'requirements': []
})

for row in ws_puestos.iter_rows(min_row=2, max_row=ws_puestos.max_row, values_only=True):
    area_name = row[0].strip() if row[0] else None
    categoria = row[1].strip() if row[1] else None
    puesto_name = row[2].strip() if row[2] else None
    mision = row[3].strip() if row[3] else ""
    especialidad = row[4].strip() if row[4] else ""
    exp_general = row[6]
    carrera_afin = row[9].strip() if row[9] else ""
    idioma = row[10].strip() if row[10] else ""

    if not area_name or not puesto_name:
        continue

    areas_data[area_name]['positions'].append({
        'title': puesto_name,
        'functions': funciones_dict.get(puesto_name, []),
        'specialty': especialidad,
        'experience': exp_general,
        'career': carrera_afin,
        'language': idioma
    })
    
    # Agregar funciones a la lista del área
    areas_data[area_name]['all_functions'].extend(funciones_dict.get(puesto_name, []))
    
    if especialidad:
        areas_data[area_name]['requirements'].append(especialidad)
    if carrera_afin:
        areas_data[area_name]['requirements'].append(carrera_afin)

def generate_choice_question(area_name, functions_list, question_num):
    """Genera una pregunta de opción múltiple basada en funciones"""
    if len(functions_list) < 2:
        return None
    
    # Seleccionar 3-4 funciones reales
    real_functions = random.sample(functions_list, min(3, len(functions_list)))
    
    # Crear una función falsa plausible
    false_options = [
        f"Reportar frecuentemente al CEO",
        f"Realizar auditorías externas",
        f"Diseñar campañas de marketing global",
        f"Gestionar presupuestos internacionales",
        f"Entrenar personal del área de ventas",
        f"Coordinar con proveedores internacionales",
        f"Desarrollar software interno",
    ]
    
    false_function = random.choice(false_options)
    
    # Construir opciones
    options = real_functions + [false_function]
    random.shuffle(options)
    
    # Encontrar el índice de la opción falsa (respuesta correcta)
    answer_index = options.index(false_function)
    
    question = {
        "type": "choice",
        "question": f"De las siguientes opciones, ¿cuál NO es una función típica en el área de {area_name}?",
        "options": options,
        "answer": answer_index
    }
    
    return question

def generate_boolean_question(area_name, requirements):
    """Genera una pregunta Verdadero/Falso"""
    if not requirements:
        return None
    
    req = random.choice(requirements)
    
    if random.choice([True, False]):
        # Verdadero
        question = {
            "type": "boolean",
            "question": f"¿Es importante tener experiencia en '{req}' para trabajar en {area_name}?",
            "answer": True
        }
    else:
        # Falso - pregunta invertida
        fake_req = random.choice([
            "Administración de Redes",
            "Cirugía Avanzada",
            "Diseño de Moda",
            "Composición Musical",
            "Pilotaje de Aviones"
        ])
        question = {
            "type": "boolean",
            "question": f"¿Es importante tener experiencia en '{fake_req}' para trabajar en {area_name}?",
            "answer": False
        }
    
    return question

def generate_experience_question(area_name):
    """Genera una pregunta sobre experiencia"""
    question = {
        "type": "choice",
        "question": f"¿Cuál es el aspecto más importante para trabajar en el área de {area_name}?",
        "options": [
            "Experiencia en el sector y especialización técnica",
            "Tener muchos contactos en redes sociales",
            "Haber estudiado en universidades internacionales",
            "Conocer idiomas exóticos"
        ],
        "answer": 0  # Primera opción
    }
    return question

def generate_quiz_for_area(area_name, area_info):
    """Genera 3-5 preguntas para un área"""
    quiz = []
    
    # Preguntas basadas en funciones (máximo 2)
    if area_info['all_functions']:
        for _ in range(min(2, len(area_info['all_functions']) // 5)):
            q = generate_choice_question(area_name, area_info['all_functions'], len(quiz))
            if q:
                quiz.append(q)
    
    # Pregunta sobre requisitos (máximo 1)
    if area_info['requirements']:
        q = generate_boolean_question(area_name, area_info['requirements'])
        if q:
            quiz.append(q)
    
    # Pregunta sobre experiencia/importancia
    q = generate_experience_question(area_name)
    if q:
        quiz.append(q)
    
    # Asegurar mínimo 3 preguntas
    while len(quiz) < 3 and len(quiz) < 5:
        if random.choice([True, False]):
            q = generate_boolean_question(area_name, area_info['requirements'])
        else:
            q = generate_choice_question(area_name, area_info['all_functions'], len(quiz))
        if q:
            quiz.append(q)
    
    # Limitar a máximo 5 preguntas
    return quiz[:5]

# Generar quiz para cada área
quiz_by_area = {}
for area_name in sorted(areas_data.keys()):
    area_info = areas_data[area_name]
    if area_info['all_functions']:  # Solo si hay funciones
        quiz_by_area[slugify(area_name)] = generate_quiz_for_area(area_name, area_info)
        print(f"✓ Generadas {len(quiz_by_area[slugify(area_name)])} preguntas para {area_name}")

# Generar archivo JavaScript
js_output = f"""(function (global) {{
  // Quiz de Career Quest
  // Generado automáticamente desde Excel
  
  const quizzesByArea = {json.dumps(quiz_by_area, ensure_ascii=False, indent=2)};

  global.quizzesByArea = quizzesByArea;

  if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ quizzesByArea }};
  }}
}})(typeof window !== 'undefined' ? window : global);
"""

# Guardar archivo
with open('js/quizData.js', 'w', encoding='utf-8') as f:
    f.write(js_output)

print(f"\n✅ Archivo quizData.js generado con éxito")
print(f"Total de áreas con quizzes: {len(quiz_by_area)}")
print(f"Total de preguntas: {sum(len(q) for q in quiz_by_area.values())}")
