#!/usr/bin/env python3
"""
Conversor de Excel a JavaScript para Career Quest
Extrae datos de "Lista de Puestos" y genera areasData.js con categorías
"""

import openpyxl
import json
from collections import defaultdict
import re
import unicodedata

def slugify(text):
    """Convierte texto a slug (lowercase con hyphens)"""
    text = unicodedata.normalize('NFKD', str(text))
    text = text.encode('ASCII', 'ignore').decode('ASCII')
    text = re.sub(r'[^\w\s-]', '', text).strip()
    text = re.sub(r'[-\s]+', '-', text)
    return text.lower()

# Colores para áreas
COLORS = [
    "#3B82F6",  # Blue
    "#8B5CF6",  # Purple
    "#EC4899",  # Pink
    "#F59E0B",  # Amber
    "#10B981",  # Green
    "#06B6D4",  # Cyan
    "#EF4444",  # Red
]

# Iconos para áreas
AREA_ICONS = {
    "almacén": "🏢",
    "comercial": "💼",
    "logística": "🚚",
    "operaciones": "⚙️",
    "contratos": "📋",
    "planeamiento": "📊",
    "supply": "📦",
}

def get_icon_for_area(area_name):
    """Retorna un ícono para el área"""
    name_lower = area_name.lower()
    for key, icon in AREA_ICONS.items():
        if key in name_lower:
            return icon
    return "🏢"

# Cargar Excel
wb = openpyxl.load_workbook('js/EXCEL/Informacion - Puestos LyC VF.xlsx')
ws_puestos = wb['Lista de Puestos']
ws_funciones = wb['Funciones']

# Extraer funciones del Excel
funciones_dict = defaultdict(list)
for row in ws_funciones.iter_rows(min_row=2, max_row=ws_funciones.max_row, values_only=True):
    if row[0] and row[1]:
        puesto_name = row[0].strip()
        funcion = row[1].strip()
        if funcion:
            funciones_dict[puesto_name].append(funcion)

# Procesar puestos por área y categoría
areas_by_name = {}
color_idx = 0

for row in ws_puestos.iter_rows(min_row=2, max_row=ws_puestos.max_row, values_only=True):
    area_name = row[0].strip() if row[0] else None
    categoria = row[1].strip() if row[1] else None
    puesto_name = row[2].strip() if row[2] else None
    mision = row[3].strip() if row[3] else ""
    especialidad = row[4].strip() if row[4] else ""
    exp_general = row[6]
    exp_puesto = row[7]
    exp_sector = row[8]
    carrera_afin = row[9].strip() if row[9] else ""
    idioma = row[10].strip() if row[10] else ""
    nivel = row[11].strip() if row[11] else ""

    if not area_name or not puesto_name:
        continue

    # Crear entrada de área si no existe
    if area_name not in areas_by_name:
        area_id = slugify(area_name)
        areas_by_name[area_name] = {
            "id": area_id,
            "title": area_name,
            "icon": get_icon_for_area(area_name),
            "summary": f"Área de {area_name}",
            "accent": COLORS[color_idx % len(COLORS)],
            "color": COLORS[color_idx % len(COLORS)],
            "departmentTitle": f"Gerencia de {area_name}",
            "description": mision,
            "npc": "Gerente",
            "npcRole": f"Responsable de {area_name}",
            "quote": f"Bienvenido al área de {area_name}",
            "competencies": [],
            "mapLabel": area_name.split()[0],  # Primera palabra
            "positions": [],
            "functions": [],
            "x": 100 + (color_idx * 80) % 600,
            "y": 100 + (color_idx * 60) % 400
        }
        color_idx += 1

    # Crear posición
    position = {
        "title": puesto_name,
        "category": categoria,
        "level": f"Nivel {int(exp_general) if exp_general else 1}",
        "blurb": especialidad,
        "speciality": especialidad,
        "experience_general": exp_general,
        "experience_position": exp_puesto,
        "experience_sector": exp_sector,
        "carrera_afin": carrera_afin,
        "idioma": idioma,
        "nivel": nivel,
        "functions": funciones_dict.get(puesto_name, [])
    }

    areas_by_name[area_name]["positions"].append(position)

# Generar areasData.js
areas_list = list(areas_by_name.values())
js_output = f"""(function (global) {{
  const areaCatalog = {json.dumps(areas_list, ensure_ascii=False, indent=2)};

  const roleOptions = areaCatalog;
  const departments = areaCatalog;

  global.areaCatalog = areaCatalog;
  global.roleOptions = roleOptions;
  global.departments = departments;

  if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ areaCatalog, roleOptions, departments }};
  }}
}})(typeof window !== 'undefined' ? window : global);
"""

# Guardar archivo
with open('js/areasData.js', 'w', encoding='utf-8') as f:
    f.write(js_output)

# También generar JSON para referencia
json_output = {
    "areas": areas_list,
    "total_positions": sum(len(a["positions"]) for a in areas_list),
    "total_functions": sum(len(a["functions"]) for a in areas_list)
}

with open('js/areasData.json', 'w', encoding='utf-8') as f:
    json.dump(json_output, f, ensure_ascii=False, indent=2)

print(f"✅ Datos convertidos exitosamente")
print(f"   - Áreas: {len(areas_list)}")
print(f"   - Puestos totales: {json_output['total_positions']}")
print(f"   - Funciones totales: {json_output['total_functions']}")
print(f"\nArchivos generados:")
print(f"   - js/areasData.js")
print(f"   - js/areasData.json")
