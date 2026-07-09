#!/usr/bin/env python3
import openpyxl
import json
from collections import defaultdict

# Cargar el Excel
wb = openpyxl.load_workbook('js/EXCEL/Informacion - Puestos LyC VF.xlsx')
ws = wb['Lista de Puestos']

# Extraer categorías y sus áreas asociadas
categories_data = defaultdict(set)
categories_order = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    area = row[0]
    categoria = row[1]
    if categoria and area:
        categoria = categoria.strip()
        area = area.strip()
        categories_data[categoria].add(area)
        if categoria not in categories_order:
            categories_order.append(categoria)

# Definir íconos y colores para categorías
category_icons = {
    'Analista': '📊',
    'Asistente': '🤝',
    'Auxiliar': '⚙️',
    'Jefe': '👔',
    'Superintendente': '🏆',
    'Supervisor': '👨‍💼'
}

category_colors = {
    'Analista': '#3B82F6',       # Blue
    'Asistente': '#8B5CF6',      # Purple
    'Auxiliar': '#EC4899',       # Pink
    'Jefe': '#F59E0B',           # Amber
    'Superintendente': '#EF4444', # Red
    'Supervisor': '#10B981'      # Green
}

category_descriptions = {
    'Analista': 'Especialista en análisis y evaluación de procesos',
    'Asistente': 'Apoyo administrativo y operativo',
    'Auxiliar': 'Personal de soporte operativo',
    'Jefe': 'Liderazgo de equipos y áreas',
    'Superintendente': 'Dirección estratégica de operaciones',
    'Supervisor': 'Supervisión directa de equipos'
}

# Crear estructura de datos
categories_list = []
for idx, cat in enumerate(categories_order):
    categories_list.append({
        'id': cat.lower().replace(' ', '-'),
        'title': cat,
        'icon': category_icons.get(cat, '💼'),
        'description': category_descriptions.get(cat, f'Categoría: {cat}'),
        'color': category_colors.get(cat, '#6366F1'),
        'areas': sorted(list(categories_data[cat]))
    })

# Generar archivo JavaScript
js_content = f"""// Categorías de puestos extraídas del Excel
const positionCategories = {json.dumps(categories_list, ensure_ascii=False, indent=2)};

// Exportar para uso en el juego
if (typeof window !== 'undefined') {{
  window.positionCategories = positionCategories;
}}
if (typeof module !== 'undefined' && module.exports) {{
  module.exports = {{ positionCategories }};
}}
"""

# Guardar archivo
with open('js/categoriesData.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print("✅ Categorías extraídas y guardadas en js/categoriesData.js")
print(f"\nCategorías encontradas ({len(categories_list)}):")
for cat in categories_list:
    print(f"  {cat['icon']} {cat['title']}: {len(cat['areas'])} áreas")
